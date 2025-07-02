from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QFrame, QGridLayout, QScrollArea,
                               QProgressBar, QListWidget, QListWidgetItem, QDateEdit, QDialogButtonBox, QDialog,
                               QMessageBox, QFileDialog, QProgressDialog, QCheckBox, QSpinBox)
from PySide6.QtCore import Qt, QTimer, QDate, Signal, QThread
from PySide6.QtGui import QFont, QPixmap, QPainter, QBrush
from modules.helper_functions import (to_persian_number, show_error_message_box, show_warning_message_box,
                                      show_information_message_box)

import sqlite3
import os
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import jdatetime
import shutil


class BackupThread(QThread):
    """Thread for database backup operations"""
    progress = Signal(int)
    status = Signal(str)
    finished_signal = Signal(bool, str)

    def __init__(self, source_db, backup_path):
        super().__init__()
        self.source_db = source_db
        self.backup_path = backup_path

    def run(self):
        try:
            self.status.emit("در حال شروع پشتیبان‌گیری...")
            self.progress.emit(10)

            # Create backup directory if it doesn't exist
            backup_dir = os.path.dirname(self.backup_path)
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)

            self.status.emit("در حال کپی کردن پایگاه داده...")
            self.progress.emit(30)

            # Copy database file
            shutil.copy2(self.source_db, self.backup_path)

            self.status.emit("در حال بررسی یکپارچگی پشتیبان...")
            self.progress.emit(70)

            # Verify backup integrity
            with sqlite3.connect(self.backup_path) as conn:
                cursor = conn.cursor()
                cursor.execute("PRAGMA integrity_check")
                result = cursor.fetchone()
                if result[0] != "ok":
                    raise Exception("خطا در یکپارچگی فایل پشتیبان")

            self.progress.emit(100)
            self.status.emit("پشتیبان‌گیری با موفقیت تکمیل شد")
            self.finished_signal.emit(True, "پشتیبان‌گیری با موفقیت انجام شد")

        except Exception as e:
            self.finished_signal.emit(False, f"خطا در پشتیبان‌گیری: {str(e)}")


class CleanupThread(QThread):
    """Thread for system cleanup operations"""
    progress = Signal(int)
    status = Signal(str)
    finished_signal = Signal(bool, str, dict)

    def __init__(self, db_path, options):
        super().__init__()
        self.db_path = db_path
        self.options = options
        self.cleanup_stats = {
            'old_logs_deleted': 0,
            'failed_logins_deleted': 0,
            'temp_files_deleted': 0,
            'space_freed': 0
        }

    def run(self):
        try:
            total_steps = sum(1 for option in self.options.values() if option)
            current_step = 0

            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                if self.options.get('old_logs', False):
                    current_step += 1
                    self.status.emit("در حال پاک کردن لاگ‌های قدیمی...")
                    self.progress.emit(int((current_step / total_steps) * 30))
                    self.cleanup_old_logs(cursor)

                if self.options.get('failed_logins', False):
                    current_step += 1
                    self.status.emit("در حال پاک کردن ورودهای ناموفق...")
                    self.progress.emit(int((current_step / total_steps) * 60))
                    self.cleanup_failed_logins(cursor)

                if self.options.get('optimize_db', False):
                    current_step += 1
                    self.status.emit("در حال بهینه‌سازی پایگاه داده...")
                    self.progress.emit(int((current_step / total_steps) * 80))
                    self.optimize_database(cursor)

                if self.options.get('temp_files', False):
                    current_step += 1
                    self.status.emit("در حال پاک کردن فایل‌های موقت...")
                    self.progress.emit(int((current_step / total_steps) * 90))
                    self.cleanup_temp_files()

                conn.commit()

            self.progress.emit(100)
            self.status.emit("پاکسازی سیستم تکمیل شد")
            self.finished_signal.emit(True, "پاکسازی سیستم با موفقیت انجام شد", self.cleanup_stats)

        except Exception as e:
            self.finished_signal.emit(False, f"خطا در پاکسازی: {str(e)}", self.cleanup_stats)

    def cleanup_old_logs(self, cursor):
        """Clean up old log entries"""
        cutoff_date = (datetime.now() - timedelta(days=self.options.get('log_days', 30))).strftime('%Y-%m-%d')

        # Count old logs before deletion
        cursor.execute("SELECT COUNT(*) FROM login_logs WHERE login_time < ?", (cutoff_date,))
        old_logs_count = cursor.fetchone()[0]

        # Delete old logs
        cursor.execute("DELETE FROM login_logs WHERE login_time < ?", (cutoff_date,))

        self.cleanup_stats['old_logs_deleted'] = old_logs_count

    def cleanup_failed_logins(self, cursor):
        """Clean up failed login attempts"""
        cutoff_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

        # Count failed logins before deletion
        cursor.execute("SELECT COUNT(*) FROM login_logs WHERE status = 'failed' AND login_time < ?", (cutoff_date,))
        failed_count = cursor.fetchone()[0]

        # Delete old failed logins
        cursor.execute("DELETE FROM login_logs WHERE status = 'failed' AND login_time < ?", (cutoff_date,))

        self.cleanup_stats['failed_logins_deleted'] = failed_count

    def optimize_database(self, cursor):
        """Optimize database"""
        cursor.execute("VACUUM")
        cursor.execute("ANALYZE")

    def cleanup_temp_files(self):
        """Clean up temporary files"""
        temp_dirs = [
            os.path.join(os.getcwd(), 'temp'),
            os.path.join(os.getcwd(), 'cache'),
            os.path.join(os.getcwd(), 'logs', 'temp')
        ]

        files_deleted = 0
        space_freed = 0

        for temp_dir in temp_dirs:
            if os.path.exists(temp_dir):
                for filename in os.listdir(temp_dir):
                    file_path = os.path.join(temp_dir, filename)
                    if os.path.isfile(file_path):
                        try:
                            file_size = os.path.getsize(file_path)
                            os.remove(file_path)
                            files_deleted += 1
                            space_freed += file_size
                        except OSError:
                            pass

        self.cleanup_stats['temp_files_deleted'] = files_deleted
        self.cleanup_stats['space_freed'] = space_freed


def backup_database(self):
    """Backup database with progress dialog"""
    try:
        # Get backup location
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"database_backup_{timestamp}.db"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "انتخاب مکان ذخیره پشتیبان",
            default_filename,
            "Database Files (*.db);;All Files (*)"
        )

        if not file_path:
            return

        # Create progress dialog
        progress_dialog = QProgressDialog("در حال پشتیبان‌گیری از پایگاه داده...", "لغو", 0, 100, self)
        progress_dialog.setWindowTitle("پشتیبان‌گیری")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setAutoClose(False)
        progress_dialog.setAutoReset(False)

        # Create status label
        status_label = QLabel("آماده‌سازی...")
        progress_dialog.setLabel(status_label)

        # Create and start backup thread
        backup_thread = BackupThread(self.db_path, file_path)

        def update_progress(value):
            progress_dialog.setValue(value)

        def update_status(status):
            status_label.setText(status)

        def backup_finished(success, message):
            progress_dialog.close()
            if success:
                QMessageBox.information(self, "موفقیت", message + f"\nمکان فایل: {file_path}")
            else:
                QMessageBox.critical(self, "خطا", message)
            backup_thread.quit()
            backup_thread.wait()

        def cancel_backup():
            backup_thread.terminate()
            backup_thread.wait()
            QMessageBox.information(self, "لغو شد", "عملیات پشتیبان‌گیری لغو شد")

        backup_thread.progress.connect(update_progress)
        backup_thread.status.connect(update_status)
        backup_thread.finished_signal.connect(backup_finished)
        progress_dialog.canceled.connect(cancel_backup)

        backup_thread.start()
        progress_dialog.show()

    except Exception as e:
        QMessageBox.critical(self, "خطا", f"خطا در شروع پشتیبان‌گیری: {str(e)}")


def cleanup_system(self):
    """System cleanup with options dialog"""
    try:
        # Create cleanup options dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("تنظیمات پاکسازی سیستم")
        dialog.setFixedSize(400, 350)
        dialog.setModal(True)

        layout = QVBoxLayout(dialog)

        # Title
        title_label = QLabel("انتخاب گزینه‌های پاکسازی:")
        title_label.setStyleSheet("font-weight: bold; font-size: 14px; margin: 10px;")
        layout.addWidget(title_label)

        # Cleanup options
        old_logs_cb = QCheckBox("پاک کردن لاگ‌های قدیمی")
        old_logs_cb.setChecked(True)
        layout.addWidget(old_logs_cb)

        # Log retention days
        log_days_layout = QHBoxLayout()
        log_days_layout.addWidget(QLabel("نگهداری لاگ‌ها برای (روز):"))
        log_days_spin = QSpinBox()
        log_days_spin.setRange(1, 365)
        log_days_spin.setValue(30)
        log_days_spin.setEnabled(old_logs_cb.isChecked())
        log_days_layout.addWidget(log_days_spin)
        log_days_layout.addStretch()
        layout.addLayout(log_days_layout)

        # Connect checkbox to spinbox
        old_logs_cb.toggled.connect(log_days_spin.setEnabled)

        failed_logins_cb = QCheckBox("پاک کردن ورودهای ناموفق قدیمی (۷ روز گذشته)")
        failed_logins_cb.setChecked(True)
        layout.addWidget(failed_logins_cb)

        optimize_db_cb = QCheckBox("بهینه‌سازی پایگاه داده")
        optimize_db_cb.setChecked(True)
        layout.addWidget(optimize_db_cb)

        temp_files_cb = QCheckBox("پاک کردن فایل‌های موقت")
        temp_files_cb.setChecked(True)
        layout.addWidget(temp_files_cb)

        # Warning
        warning_label = QLabel("⚠️ توجه: این عملیات قابل برگشت نیست. قبل از ادامه از پایگاه داده پشتیبان تهیه کنید.")
        warning_label.setStyleSheet(
            "color: #e74c3c; margin: 10px; padding: 10px; background-color: #fdf2f2; border: 1px solid #e74c3c; border-radius: 5px;")
        warning_label.setWordWrap(True)
        layout.addWidget(warning_label)

        # Buttons
        button_layout = QHBoxLayout()

        backup_first_btn = QPushButton("ابتدا پشتیبان بگیر")
        backup_first_btn.setStyleSheet("background-color: #f39c12; color: white; padding: 8px; border-radius: 4px;")
        backup_first_btn.clicked.connect(lambda: (dialog.accept(), self.backup_database()))

        proceed_btn = QPushButton("ادامه پاکسازی")
        proceed_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px; border-radius: 4px;")

        cancel_btn = QPushButton("لغو")
        cancel_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px; border-radius: 4px;")
        cancel_btn.clicked.connect(dialog.reject)

        button_layout.addWidget(backup_first_btn)
        button_layout.addWidget(proceed_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)

        def proceed_cleanup():
            dialog.accept()

            # Gather cleanup options
            cleanup_options = {
                'old_logs': old_logs_cb.isChecked(),
                'log_days': log_days_spin.value(),
                'failed_logins': failed_logins_cb.isChecked(),
                'optimize_db': optimize_db_cb.isChecked(),
                'temp_files': temp_files_cb.isChecked()
            }

            # Check if any option is selected
            if not any(cleanup_options.values()):
                QMessageBox.information(self, "اطلاع", "هیچ گزینه‌ای انتخاب نشده است.")
                return

            self.perform_cleanup(cleanup_options)

        proceed_btn.clicked.connect(proceed_cleanup)

        dialog.exec_()

    except Exception as e:
        QMessageBox.critical(self, "خطا", f"خطا در تنظیمات پاکسازی: {str(e)}")


def perform_cleanup(self, cleanup_options):
    """Perform the actual cleanup with progress dialog"""
    try:
        # Create progress dialog
        progress_dialog = QProgressDialog("در حال پاکسازی سیستم...", "لغو", 0, 100, self)
        progress_dialog.setWindowTitle("پاکسازی سیستم")
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setAutoClose(False)
        progress_dialog.setAutoReset(False)

        # Create status label
        status_label = QLabel("آماده‌سازی...")
        progress_dialog.setLabel(status_label)

        # Create and start cleanup thread
        cleanup_thread = CleanupThread(self.db_path, cleanup_options)

        def update_progress(value):
            progress_dialog.setValue(value)

        def update_status(status):
            status_label.setText(status)

        def cleanup_finished(success, message, stats):
            progress_dialog.close()

            if success:
                # Show detailed results
                results_message = message + "\n\nجزئیات:\n"
                if stats['old_logs_deleted'] > 0:
                    results_message += f"• {stats['old_logs_deleted']} لاگ قدیمی پاک شد\n"
                if stats['failed_logins_deleted'] > 0:
                    results_message += f"• {stats['failed_logins_deleted']} ورود ناموفق پاک شد\n"
                if stats['temp_files_deleted'] > 0:
                    results_message += f"• {stats['temp_files_deleted']} فایل موقت پاک شد\n"
                if stats['space_freed'] > 0:
                    space_mb = stats['space_freed'] / (1024 * 1024)
                    results_message += f"• {space_mb:.2f} مگابایت فضا آزاد شد\n"

                QMessageBox.information(self, "موفقیت", results_message)

                # Refresh dashboard after cleanup
                self.refresh_dashboard()
            else:
                QMessageBox.critical(self, "خطا", message)

            cleanup_thread.quit()
            cleanup_thread.wait()

        def cancel_cleanup():
            cleanup_thread.terminate()
            cleanup_thread.wait()
            QMessageBox.information(self, "لغو شد", "عملیات پاکسازی لغو شد")

        cleanup_thread.progress.connect(update_progress)
        cleanup_thread.status.connect(update_status)
        cleanup_thread.finished_signal.connect(cleanup_finished)
        progress_dialog.canceled.connect(cancel_cleanup)

        cleanup_thread.start()
        progress_dialog.show()

    except Exception as e:
        QMessageBox.critical(self, "خطا", f"خطا در شروع پاکسازی: {str(e)}")


class AdminDashboardWidget(QWidget):
    """Main admin dashboard widget with statistics and system monitoring"""

    def __init__(self, parent, username, db_path):
        super().__init__()
        self.username = username
        self.db_path = db_path
        self.user_data = None
        self.setup_ui()
        self.load_admin_data()
        self.setup_dashboard_timer()

    def setup_ui(self):
        """Setup the dashboard user interface"""
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f6fa;
                font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
            }
            QFrame#info_frame {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 8px;
                padding: 15px;
                margin: 5px;
            }
            QFrame#dashboard_card {
                background-color: white;
                border: 1px solid #e1e8ed;
                border-radius: 10px;
                padding: 20px;
                margin: 10px;
            }
            QLabel#title_label {
                color: #2c3e50;
                font-size: 18px;
                font-weight: bold;
                padding: 5px;
            }
            QLabel#subtitle_label {
                color: #7f8c8d;
                font-size: 14px;
                padding: 2px;
            }
            QLabel#value_label {
                color: #27ae60;
                font-size: 24px;
                font-weight: bold;
                padding: 5px;
            }
            QLabel#warning_label {
                color: #e74c3c;
                font-size: 24px;
                font-weight: bold;
                padding: 5px;
            }
            QPushButton#dashboard_btn {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton#dashboard_btn:hover {
                background-color: #2980b9;
            }
            QPushButton#logout_btn {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton#logout_btn:hover {
                background-color: #c0392b;
            }
            QProgressBar {
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                text-align: center;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 3px;
            }
            QListWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                background-color: #fafafa;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #eee;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
        """)

        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header section
        header_frame = QFrame()
        header_frame.setObjectName("info_frame")
        header_layout = QHBoxLayout(header_frame)

        # Profile picture placeholder
        self.profile_pic = QLabel()
        self.profile_pic.setFixedSize(100, 100)
        self.profile_pic.setStyleSheet("""
            border: 2px solid #3498db;
            border-radius: 50px;
            background-color: #ecf0f1;
        """)
        self.profile_pic.setAlignment(Qt.AlignCenter)
        self.profile_pic.setText("👤")
        self.profile_pic.setFont(QFont("", 36))

        # Admin info
        info_layout = QVBoxLayout()
        self.admin_name_label = QLabel("نام مدیر")
        self.admin_name_label.setObjectName("title_label")
        self.admin_role_label = QLabel("نقش")
        self.admin_role_label.setObjectName("subtitle_label")

        info_layout.addWidget(self.admin_name_label)
        info_layout.addWidget(self.admin_role_label)
        info_layout.addStretch()

        header_layout.addWidget(self.profile_pic)
        header_layout.addLayout(info_layout)
        header_layout.addStretch()

        main_layout.addWidget(header_frame)

        # Dashboard section
        dashboard_scroll = QScrollArea()
        dashboard_widget = QWidget()
        self.dashboard_layout = QGridLayout(dashboard_widget)

        self.create_dashboard_cards()

        dashboard_scroll.setWidget(dashboard_widget)
        dashboard_scroll.setWidgetResizable(True)
        main_layout.addWidget(dashboard_scroll)

        self.setLayout(main_layout)

    def create_dashboard_cards(self):
        """Create dashboard cards"""
        # System Statistics Card
        stats_card = self.create_stats_card()
        self.dashboard_layout.addWidget(stats_card, 0, 0)

        # User Activity Card
        activity_card = self.create_activity_card()
        self.dashboard_layout.addWidget(activity_card, 0, 1)

        # System Health Card
        health_card = self.create_health_card()
        self.dashboard_layout.addWidget(health_card, 1, 0)

        # Recent Actions Card
        actions_card = self.create_actions_card()
        self.dashboard_layout.addWidget(actions_card, 1, 1)

        # Quick Actions Card
        quick_actions_card = self.create_quick_actions_card()
        self.dashboard_layout.addWidget(quick_actions_card, 2, 0, 1, 2)

    def create_stats_card(self):
        """Create system statistics card"""
        card = QFrame()
        card.setObjectName("dashboard_card")
        layout = QVBoxLayout(card)

        title = QLabel("آمار سیستم")
        title.setObjectName("title_label")
        layout.addWidget(title)

        # Stats grid
        stats_grid = QGridLayout()

        self.total_users_label = QLabel("0")
        self.total_users_label.setObjectName("value_label")
        stats_grid.addWidget(QLabel("کل کاربران:"), 0, 0)
        stats_grid.addWidget(self.total_users_label, 0, 1)

        self.active_users_label = QLabel("0")
        self.active_users_label.setObjectName("value_label")
        stats_grid.addWidget(QLabel("کاربران فعال:"), 1, 0)
        stats_grid.addWidget(self.active_users_label, 1, 1)

        # Changed from "کل ورودها" to "ورودهای امروز"
        self.total_logins_label = QLabel("0")
        self.total_logins_label.setObjectName("value_label")
        stats_grid.addWidget(QLabel("ورودهای امروز:"), 2, 0)
        stats_grid.addWidget(self.total_logins_label, 2, 1)

        layout.addLayout(stats_grid)
        layout.addStretch()

        return card

    def create_activity_card(self):
        """Create user activity card"""
        card = QFrame()
        card.setObjectName("dashboard_card")
        layout = QVBoxLayout(card)

        title = QLabel("فعالیت کاربران")
        title.setObjectName("title_label")
        layout.addWidget(title)

        # Activity progress bars
        self.admin_activity = QProgressBar()
        self.admin_activity.setMaximum(100)
        layout.addWidget(QLabel("مدیران:"))
        layout.addWidget(self.admin_activity)

        self.translator_activity = QProgressBar()
        self.translator_activity.setMaximum(100)
        layout.addWidget(QLabel("مترجمان:"))
        layout.addWidget(self.translator_activity)

        self.clerk_activity = QProgressBar()
        self.clerk_activity.setMaximum(100)
        layout.addWidget(QLabel("منشی‌ها:"))
        layout.addWidget(self.clerk_activity)

        self.accountant_activity = QProgressBar()
        self.accountant_activity.setMaximum(100)
        layout.addWidget(QLabel("حسابداران:"))
        layout.addWidget(self.accountant_activity)

        layout.addStretch()

        return card

    def create_health_card(self):
        """Create system health card"""
        card = QFrame()
        card.setObjectName("dashboard_card")
        layout = QVBoxLayout(card)

        title = QLabel("وضعیت سیستم")
        title.setObjectName("title_label")
        layout.addWidget(title)

        # Health indicators
        health_grid = QGridLayout()

        self.db_status_label = QLabel("✅ متصل")
        self.db_status_label.setObjectName("value_label")
        health_grid.addWidget(QLabel("پایگاه داده:"), 0, 0)
        health_grid.addWidget(self.db_status_label, 0, 1)

        self.uptime_label = QLabel("00:00:00")
        self.uptime_label.setObjectName("value_label")
        health_grid.addWidget(QLabel("زمان فعالیت:"), 1, 0)
        health_grid.addWidget(self.uptime_label, 1, 1)

        # Changed from "ورودهای ناموفق" to "تعداد فعالیت‌ها"
        self.failed_logins_label = QLabel("0")
        self.failed_logins_label.setObjectName("value_label")
        health_grid.addWidget(QLabel("تعداد فعالیت‌ها:"), 2, 0)
        health_grid.addWidget(self.failed_logins_label, 2, 1)

        layout.addLayout(health_grid)
        layout.addStretch()

        return card

    def create_actions_card(self):
        """Create recent actions card"""
        card = QFrame()
        card.setObjectName("dashboard_card")
        layout = QVBoxLayout(card)

        title = QLabel("فعالیت‌های اخیر")
        title.setObjectName("title_label")
        layout.addWidget(title)

        self.recent_actions_list = QListWidget()
        layout.addWidget(self.recent_actions_list)

        return card

    def create_quick_actions_card(self):
        """Create quick actions card"""
        card = QFrame()
        card.setObjectName("dashboard_card")
        layout = QVBoxLayout(card)

        title = QLabel("عملیات سریع")
        title.setObjectName("title_label")
        layout.addWidget(title)

        # Quick action buttons
        buttons_layout = QHBoxLayout()

        refresh_btn = QPushButton("بروزرسانی داشبورد")
        refresh_btn.setObjectName("dashboard_btn")
        refresh_btn.clicked.connect(self.refresh_dashboard)

        backup_btn = QPushButton("پشتیبان‌گیری")
        backup_btn.setObjectName("dashboard_btn")
        backup_btn.clicked.connect(self.backup_database)

        cleanup_btn = QPushButton("پاکسازی سیستم")
        cleanup_btn.setObjectName("dashboard_btn")
        cleanup_btn.clicked.connect(self.cleanup_system)

        export_btn = QPushButton("گزارش‌گیری")
        export_btn.setObjectName("dashboard_btn")
        export_btn.clicked.connect(self.export_reports)

        buttons_layout.addWidget(refresh_btn)
        buttons_layout.addWidget(backup_btn)
        buttons_layout.addWidget(cleanup_btn)
        buttons_layout.addWidget(export_btn)

        layout.addLayout(buttons_layout)

        return card

    def load_admin_data(self):
        """Load admin data from database"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Get admin profile data
                cursor.execute("""
                    SELECT up.full_name, up.role_fa, up.email, up.phone, up.avatar_path,
                           u.role, u.start_date
                    FROM user_profiles up
                    JOIN users u ON up.username = u.username
                    WHERE up.username = ?
                """, (self.username,))

                result = cursor.fetchone()
                if result:
                    full_name, role_fa, email, phone, avatar_path, role, start_date = result
                    self.user_data = {
                        'full_name': full_name,
                        'role_fa': role_fa,
                        'email': email,
                        'phone': phone,
                        'avatar_path': avatar_path,
                        'role': role,
                        'start_date': start_date
                    }

                    # Update UI
                    self.admin_name_label.setText(full_name or self.username)
                    self.admin_role_label.setText(role_fa or role)

                    # Load avatar if exists
                    self.load_avatar(avatar_path)

        except sqlite3.Error as e:
            title = "خطا در بارگذاری اطلاعات مدیر"
            message = (f"خطا در بارگذاری اطلاعات مدیر: \n"
                       f"{str(e)}")
            show_error_message_box(self, title, message)

    def load_avatar(self, avatar_path):
        """Load user avatar"""
        if avatar_path and os.path.exists(avatar_path):
            try:
                pixmap = QPixmap(avatar_path)
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.profile_pic.setPixmap(self.create_circular_pixmap(scaled_pixmap))
                    self.profile_pic.setText("")
            except Exception as e:
                title = "خطا در بارگذاری آواتار"
                message = (f"خطا در بارگذاری آواتار: \n"
                            f"{str(e)}")
                show_error_message_box(self, title, message)

    def create_circular_pixmap(self, pixmap):
        """Create circular pixmap"""
        size = min(pixmap.width(), pixmap.height())
        circular_pixmap = QPixmap(size, size)
        circular_pixmap.fill(Qt.transparent)

        painter = QPainter(circular_pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setBrush(QBrush(pixmap))
        painter.setPen(Qt.NoPen)
        painter.drawEllipse(0, 0, size, size)
        painter.end()

        return circular_pixmap

    def setup_dashboard_timer(self):
        """Setup timer for dashboard updates"""
        self.dashboard_timer = QTimer()
        self.dashboard_timer.timeout.connect(self.update_dashboard)
        self.dashboard_timer.start(30000)  # Update every 30 seconds

        # Initial update
        QTimer.singleShot(1000, self.update_dashboard)

    def update_dashboard(self):
        """Update dashboard data"""
        self.update_system_stats()
        self.update_user_activity()
        self.update_system_health()
        self.update_recent_actions()

    def update_system_stats(self):
        """Update system statistics"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Total users
                cursor.execute("SELECT COUNT(*) FROM users")
                total_users = cursor.fetchone()[0]
                self.total_users_label.setText(str(total_users))

                # Active users
                cursor.execute("SELECT COUNT(*) FROM users WHERE active = 1")
                active_users = cursor.fetchone()[0]
                self.active_users_label.setText(str(active_users))

                # Today's logins (changed from total logins)
                today = datetime.now().strftime('%Y-%m-%d')
                cursor.execute("""
                    SELECT COUNT(*) FROM login_logs 
                    WHERE status = 'success' AND DATE(login_time) = ?
                """, (today,))
                today_logins = cursor.fetchone()[0]
                self.total_logins_label.setText(str(today_logins))

        except sqlite3.Error as e:
            title = "خطا در بروزرسانی آمار سیستم"
            message = (f"خطا در بروزرسانی آمار سیستم: \n"
                          f"{str(e)}")
            show_error_message_box(self, title, message)

    def update_user_activity(self):
        """Update user activity progress bars"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Get activity by role (last 7 days)
                roles = ['admin', 'translator', 'clerk', 'accountant']
                progress_bars = [self.admin_activity, self.translator_activity,
                                 self.clerk_activity, self.accountant_activity]

                for role, progress_bar in zip(roles, progress_bars):
                    cursor.execute("""
                        SELECT COUNT(DISTINCT ll.username)
                        FROM login_logs ll
                        JOIN users u ON ll.username = u.username
                        WHERE u.role = ? AND ll.login_time >= date('now', '-7 days')
                        AND ll.status = 'success'
                    """, (role,))

                    active_count = cursor.fetchone()[0]

                    cursor.execute("SELECT COUNT(*) FROM users WHERE role = ?", (role,))
                    total_count = cursor.fetchone()[0]

                    if total_count > 0:
                        percentage = int((active_count / total_count) * 100)
                        progress_bar.setValue(percentage)
                    else:
                        progress_bar.setValue(0)

        except sqlite3.Error as e:
            title = "خطا در بروزرسانی فعالیت کاربران"
            message = (f"خطا در بروزرسانی فعالیت کاربران: \n"
                            f"{str(e)}")
            show_error_message_box(self, title, message)

    def update_system_health(self):
        """Update system health indicators"""
        try:
            # Database status
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT 1")
                self.db_status_label.setText("✅ متصل")
                self.db_status_label.setObjectName("value_label")

        except sqlite3.Error:
            self.db_status_label.setText("❌ قطع")
            self.db_status_label.setObjectName("warning_label")

        # Update uptime (simplified)
        if not hasattr(self, 'start_time'):
            self.start_time = datetime.now()

        uptime = datetime.now() - self.start_time
        uptime_str = str(uptime).split('.')[0]  # Remove microseconds
        self.uptime_label.setText(uptime_str)

        # Update activity count (count of items in recent actions list)
        activity_count = self.recent_actions_list.count()
        self.failed_logins_label.setText(str(activity_count))

    def update_recent_actions(self):
        """Update recent actions list with enhanced activity tracking"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Get today's date
                today = datetime.now().strftime('%Y-%m-%d')

                self.recent_actions_list.clear()
                actions = []

                # 1. Get login/logout activities for today
                cursor.execute("""
                    SELECT ll.username, ll.login_time, ll.status, up.full_name
                    FROM login_logs ll
                    LEFT JOIN user_profiles up ON ll.username = up.username
                    WHERE DATE(ll.login_time) = ?
                    ORDER BY ll.login_time DESC
                """, (today,))

                login_activities = cursor.fetchall()
                for username, login_time, status, full_name in login_activities:
                    time_obj = datetime.fromisoformat(login_time)
                    formatted_time = time_obj.strftime("%H:%M")
                    persian_formatted_time = to_persian_number(formatted_time)

                    name = full_name if full_name else username

                    if status == "success":
                        action_text = f"✅ {name} وارد شد - {persian_formatted_time}"
                    else:
                        action_text = f"❌ {name} ورود ناموفق - {persian_formatted_time}"

                    actions.append((time_obj, action_text))

                # 2. Get invoice creation activities for today
                # Assuming you have an 'invoices' table with columns: id, created_by, created_at, invoice_no
                try:
                    cursor.execute("""
                        SELECT i.created_by, i.created_at, i.invoice_no, up.full_name
                        FROM invoices i
                        LEFT JOIN user_profiles up ON i.created_by = up.username
                        WHERE DATE(i.created_at) = ?
                        ORDER BY i.created_at DESC
                    """, (today,))

                    invoice_activities = cursor.fetchall()
                    for created_by, created_at, invoice_no, full_name in invoice_activities:
                        time_obj = datetime.fromisoformat(created_at)
                        formatted_time = time_obj.strftime("%H:%M")
                        persian_formatted_time = to_persian_number(formatted_time)

                        name = full_name if full_name else created_by
                        persian_invoice_no = to_persian_number(str(invoice_no))

                        action_text = f"📄 {name} فاکتور شماره {persian_invoice_no} را صادر کرد - {persian_formatted_time}"
                        actions.append((time_obj, action_text))
                except sqlite3.OperationalError:
                    # invoices table might not exist yet
                    pass

                # 3. Get invoice edit activities for today
                # Assuming you have an 'invoice_edits' table with columns: invoice_id, edited_by, edited_at, invoice_no
                try:
                    cursor.execute("""
                        SELECT ie.edited_by, ie.edited_at, ie.invoice_no, up.full_name
                        FROM invoice_edits ie
                        LEFT JOIN user_profiles up ON ie.edited_by = up.username
                        WHERE DATE(ie.edited_at) = ?
                        ORDER BY ie.edited_at DESC
                    """, (today,))

                    edit_activities = cursor.fetchall()
                    for edited_by, edited_at, invoice_no, full_name in edit_activities:
                        time_obj = datetime.fromisoformat(edited_at)
                        formatted_time = time_obj.strftime("%H:%M")
                        persian_formatted_time = to_persian_number(formatted_time)

                        name = full_name if full_name else edited_by
                        persian_invoice_no = to_persian_number(str(invoice_no))

                        action_text = f"✏️ {name} فاکتور شماره {persian_invoice_no} را ویرایش کرد - {persian_formatted_time}"
                        actions.append((time_obj, action_text))
                except sqlite3.OperationalError:
                    # invoice_edits table might not exist yet
                    pass

                # Sort all actions by time (most recent first)
                actions.sort(key=lambda x: x[0], reverse=True)

                # Add actions to list widget (limit to 20 most recent)
                for _, action_text in actions[:20]:
                    item = QListWidgetItem(action_text)
                    self.recent_actions_list.addItem(item)

        except sqlite3.Error as e:
            title = "خطا در بروزرسانی فعالیت‌های اخیر"
            message = (f"خطا در بروزرسانی فعالیت‌های اخیر: \n"
                            f"{str(e)}")
            show_error_message_box(self, title, message)

    def refresh_dashboard(self):
        """Refresh dashboard data"""
        self.update_dashboard()

    def backup_database(self):
        """Backup database (placeholder)"""
        self.backup_database = backup_database.__get__(self, AdminDashboardWidget)

    def cleanup_system(self):
        """Cleanup system (placeholder)"""
        self.cleanup_system = cleanup_system.__get__(self, AdminDashboardWidget)
        self.perform_cleanup = perform_cleanup.__get__(self, AdminDashboardWidget)

    def export_reports(self):
        """Export comprehensive user activity reports to Excel"""
        try:
            # Create date range dialog
            dialog = QDialog(self)
            dialog.setWindowTitle("انتخاب بازه زمانی گزارش")
            dialog.setFixedSize(400, 200)

            layout = QVBoxLayout(dialog)

            # Date selection
            start_date_label = QLabel("تاریخ شروع:")
            start_date_edit = QDateEdit()
            start_date_edit.setDate(QDate.currentDate().addDays(-30))
            start_date_edit.setCalendarPopup(True)

            end_date_label = QLabel("تاریخ پایان:")
            end_date_edit = QDateEdit()
            end_date_edit.setDate(QDate.currentDate())
            end_date_edit.setCalendarPopup(True)

            buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            buttons.accepted.connect(dialog.accept)
            buttons.rejected.connect(dialog.reject)

            layout.addWidget(start_date_label)
            layout.addWidget(start_date_edit)
            layout.addWidget(end_date_label)
            layout.addWidget(end_date_edit)
            layout.addWidget(buttons)

            if dialog.exec_() == QDialog.Accepted:
                start_date = start_date_edit.date().toString("yyyy-MM-dd")
                end_date = end_date_edit.date().toString("yyyy-MM-dd")

                self.generate_user_reports(start_date, end_date)

        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در تولید گزارش: {str(e)}")

    def generate_user_reports(self, start_date, end_date):
        """Generate Excel report with separate sheets for each user"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()

                # Get all users
                cursor.execute("""
                    SELECT DISTINCT u.username, up.full_name, u.role
                    FROM users u
                    LEFT JOIN user_profiles up ON u.username = up.username
                    ORDER BY up.full_name
                """)

                users = cursor.fetchall()

                if not users:
                    QMessageBox.information(self, "اطلاع", "هیچ کاربری یافت نشد.")
                    return

                # Create Excel workbook
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet

                # Create summary sheet
                summary_sheet = wb.create_sheet("خلاصه گزارش")
                self.create_summary_sheet(summary_sheet, cursor, start_date, end_date)

                # Create sheet for each user
                for username, full_name, role in users:
                    sheet_name = full_name if full_name else username
                    # Limit sheet name length for Excel compatibility
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:28] + "..."

                    user_sheet = wb.create_sheet(sheet_name)
                    self.create_user_sheet(user_sheet, cursor, username, full_name, role, start_date, end_date)

                # Save file
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "ذخیره گزارش",
                    f"User_Activity_Report_{start_date}_to_{end_date}.xlsx",
                    "Excel Files (*.xlsx)"
                )

                if file_path:
                    wb.save(file_path)
                    QMessageBox.information(self, "موفقیت", f"گزارش با موفقیت ذخیره شد:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در تولید گزارش: {str(e)}")

    def create_summary_sheet(self, sheet, cursor, start_date, end_date):
        """Create summary sheet with overall statistics"""
        try:
            # Set RTL direction for Persian text
            sheet.sheet_view.rightToLeft = True

            # Headers
            headers = ['نام کاربری', 'نام کامل', 'نقش', 'تعداد ورود موفق', 'تعداد ورود ناموفق', 'آخرین ورود', 'وضعیت']

            # Style for headers
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Add headers
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Get summary data for all users
            cursor.execute("""
                SELECT u.username, up.full_name, u.role, u.active,
                       COUNT(CASE WHEN ll.status = 'success' AND ll.login_time BETWEEN ? AND ? THEN 1 END) as successful_logins,
                       COUNT(CASE WHEN ll.status = 'failed' AND ll.login_time BETWEEN ? AND ? THEN 1 END) as failed_logins,
                       MAX(CASE WHEN ll.status = 'success' THEN ll.login_time END) as last_login
                FROM users u
                LEFT JOIN user_profiles up ON u.username = up.username
                LEFT JOIN login_logs ll ON u.username = ll.username
                GROUP BY u.username, up.full_name, u.role, u.active
                ORDER BY up.full_name
            """, (start_date, end_date, start_date, end_date))

            users_data = cursor.fetchall()

            # Add data rows
            for row_idx, (username, full_name, role, active, success_logins, failed_logins, last_login) in enumerate(
                    users_data, 2):
                sheet.cell(row=row_idx, column=1, value=username)
                sheet.cell(row=row_idx, column=2, value=full_name or "تعین نشده")
                sheet.cell(row=row_idx, column=3, value=self.get_role_persian(role))
                sheet.cell(row=row_idx, column=4, value=success_logins)
                sheet.cell(row=row_idx, column=5, value=failed_logins)

                if last_login:
                    formatted_date = datetime.fromisoformat(last_login).strftime("%Y/%m/%d %H:%M")
                    sheet.cell(row=row_idx, column=6, value=to_persian_number(formatted_date))
                else:
                    sheet.cell(row=row_idx, column=6, value="هرگز")

                status = "فعال" if active else "غیرفعال"
                sheet.cell(row=row_idx, column=7, value=status)

                # Color inactive users differently
                if not active:
                    for col in range(1, 8):
                        sheet.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6",
                                                                               fill_type="solid")

            # Adjust column widths
            column_widths = [20, 25, 15, 15, 15, 20, 10]
            for col, width in enumerate(column_widths, 1):
                sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = width

            # Add totals row
            total_row = len(users_data) + 3
            sheet.cell(row=total_row, column=1, value="مجموع:").font = Font(bold=True)
            sheet.cell(row=total_row, column=4, value=sum(user_data[4] for user_data in users_data)).font = Font(
                bold=True)
            sheet.cell(row=total_row, column=5, value=sum(user_data[5] for user_data in users_data)).font = Font(
                bold=True)

        except Exception as e:
            title = "خطا در ایجاد برگه خلاصه"
            message = (f"خطا در ایجاد برگه خلاصه: \n"
                          f"{str(e)}")
            show_error_message_box(self, title, message)

        except Exception as e:
            title = "خطا در ایجاد برگه خلاصه"
            message = (f"خطا در ایجاد برگه خلاصه: \n"
                            f"{str(e)}")
            show_error_message_box(self, title, message)

    def create_user_sheet(self, sheet, cursor, username, full_name, role, start_date, end_date):
        """Create detailed sheet for individual user"""
        try:
            # Set RTL direction for Persian text
            sheet.sheet_view.rightToLeft = True

            # User info header
            sheet.merge_cells('A1:E1')
            user_title = f"گزارش فعالیت: {full_name or username}"
            sheet.cell(row=1, column=1, value=user_title)
            sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            sheet.cell(row=1, column=1).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            # User details
            sheet.cell(row=2, column=1, value="نام کاربری:")
            sheet.cell(row=2, column=2, value=username)
            sheet.cell(row=3, column=1, value="نقش:")
            sheet.cell(row=3, column=2, value=self.get_role_persian(role))
            sheet.cell(row=4, column=1, value="بازه گزارش:")
            sheet.cell(row=4, column=2, value=f"{to_persian_number(start_date)} تا {to_persian_number(end_date)}")

            # Login activities section
            sheet.cell(row=6, column=1, value="فعالیت‌های ورود و خروج").font = Font(bold=True, size=12)

            # Login headers
            login_headers = ['تاریخ', 'زمان', 'وضعیت', 'آدرس IP', 'توضیحات']
            for col, header in enumerate(login_headers, 1):
                cell = sheet.cell(row=7, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Get login data
            cursor.execute("""
                SELECT login_time, status, ip_address, user_agent
                FROM login_logs
                WHERE username = ? AND login_time BETWEEN ? AND ?
                ORDER BY login_time DESC
            """, (username, start_date, end_date))

            login_data = cursor.fetchall()
            current_row = 8

            for login_time, status, ip_address, user_agent in login_data:
                dt = datetime.fromisoformat(login_time)
                date_str = to_persian_number(dt.strftime("%Y/%m/%d"))
                time_str = to_persian_number(dt.strftime("%H:%M:%S"))
                status_persian = "موفق" if status == "success" else "ناموفق"

                sheet.cell(row=current_row, column=1, value=date_str)
                sheet.cell(row=current_row, column=2, value=time_str)
                sheet.cell(row=current_row, column=3, value=status_persian)
                sheet.cell(row=current_row, column=4, value=ip_address or "نامشخص")
                sheet.cell(row=current_row, column=5,
                           value=user_agent[:50] + "..." if user_agent and len(user_agent) > 50 else user_agent or "")

                # Color failed logins
                if status == "failed":
                    for col in range(1, 6):
                        sheet.cell(row=current_row, column=col).fill = PatternFill(start_color="FFCCCB",
                                                                                   end_color="FFCCCB",
                                                                                   fill_type="solid")

                current_row += 1

            # Invoice activities section (if invoices table exists)
            try:
                invoice_section_row = current_row + 2
                sheet.cell(row=invoice_section_row, column=1, value="فعالیت‌های فاکتور").font = Font(bold=True, size=12)

                # Invoice headers
                invoice_headers = ['تاریخ', 'زمان', 'شماره فاکتور', 'نوع عملیات', 'توضیحات']
                for col, header in enumerate(invoice_headers, 1):
                    cell = sheet.cell(row=invoice_section_row + 1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                current_row = invoice_section_row + 2

                # Get invoice creation data
                cursor.execute("""
                    SELECT created_at, invoice_no, 'ایجاد' as operation_type, '' as description
                    FROM invoices
                    WHERE created_by = ? AND created_at BETWEEN ? AND ?
                    UNION ALL
                    SELECT edited_at, invoice_no, 'ویرایش' as operation_type, edit_reason as description
                    FROM invoice_edits
                    WHERE edited_by = ? AND edited_at BETWEEN ? AND ?
                    ORDER BY 1 DESC
                """, (username, start_date, end_date, username, start_date, end_date))

                invoice_data = cursor.fetchall()

                for activity_time, invoice_no, operation_type, description in invoice_data:
                    dt = datetime.fromisoformat(activity_time)
                    date_str = to_persian_number(dt.strftime("%Y/%m/%d"))
                    time_str = to_persian_number(dt.strftime("%H:%M:%S"))
                    invoice_no_persian = to_persian_number(str(invoice_no))

                    sheet.cell(row=current_row, column=1, value=date_str)
                    sheet.cell(row=current_row, column=2, value=time_str)
                    sheet.cell(row=current_row, column=3, value=invoice_no_persian)
                    sheet.cell(row=current_row, column=4, value=operation_type)
                    sheet.cell(row=current_row, column=5, value=description or "")

                    current_row += 1

            except sqlite3.OperationalError:
                # Invoice tables don't exist yet
                pass

            # Summary statistics
            stats_row = current_row + 2
            sheet.cell(row=stats_row, column=1, value="خلاصه آمار").font = Font(bold=True, size=12)

            # Calculate statistics
            successful_logins = sum(1 for data in login_data if data[1] == "success")
            failed_logins = sum(1 for data in login_data if data[1] == "failed")

            sheet.cell(row=stats_row + 1, column=1, value="ورودهای موفق:")
            sheet.cell(row=stats_row + 1, column=2, value=to_persian_number(str(successful_logins)))
            sheet.cell(row=stats_row + 2, column=1, value="ورودهای ناموفق:")
            sheet.cell(row=stats_row + 2, column=2, value=to_persian_number(str(failed_logins)))
            sheet.cell(row=stats_row + 3, column=1, value="کل فعالیت‌ها:")
            sheet.cell(row=stats_row + 3, column=2, value=to_persian_number(str(len(login_data))))

            # Adjust column widths
            column_widths = [15, 15, 15, 15, 30]
            for col, width in enumerate(column_widths, 1):
                sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = width

        except Exception as e:
            title = "خطا در ایجاد برگه کاربر"
            message = (f"خطا در ایجاد برگه کاربر {username}: \n"
                          f"{str(e)}")
            show_error_message_box(self, title, message)

    def get_role_persian(self, role):
        """Convert English role to Persian"""
        role_mapping = {
            'admin': 'مدیر',
            'translator': 'مترجم',
            'clerk': 'منشی',
            'accountant': 'حسابدار'
        }
        return role_mapping.get(role, role)

    def closeEvent(self, event):
        """Handle widget close event"""
        if hasattr(self, 'dashboard_timer'):
            self.dashboard_timer.stop()
        event.accept()
